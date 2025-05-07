
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2022/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

movie_rows = soup.finalAll('tr')

#print(movie_rows[1])

wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'No'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Theaters'
ws['F1'] = 'Avg Gross/Theater'

for x in range(1, 6):
    td = movie_rows[x].find_all('td')
    rank = td[0].text
    title = td[1].text
    theater = int(td[6].text.replace(",", ""))
    gross = int(td[5].text.strip('$').replace(",", ""))
    release_date = td[8].text

    avg_per_theater = round(gross / theater, 2)

    ws['A' + str(x+1)] = rank
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = release_date
    ws['C' + str(x+1)] = gross
    ws['E' + str(x+1)] = theater
    ws['F' + str(x+1)] = avg_per_theater

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 26

header_font = Font(size=16, bold=True)

for cell in ws[1]:
    cell.font = header_font

for cell in ws['E'][1:]:
    cell.number_format = '#,##0'

for cell in ws['D'][1:]:
    cell.number_format = u'"$"#,##0'

##
##
##
##
wb.save('box_office_2022.xlsx')